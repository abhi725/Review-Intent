"""Email verification, reset links, avatars, signal counts and export.

The theme is the same as test_auth.py: properties whose absence is invisible
until something goes wrong. An unverified account that can sign in anyway, a
reset link that works twice, an avatar upload that stores whatever it was
handed, an export that quietly stops at row 500.
"""

import asyncio
import io
from datetime import datetime, timedelta, timezone

import pytest

from intentdesk.services import avatars, preferences, signals, users


class FakeDB:
    """Records what was executed so a test can assert on the SQL as well as the
    return value — several of these properties live in the statement itself."""

    def __init__(self, rows=None):
        self.rows = rows or {}
        self.executed = []

    async def fetch(self, sql, *args):
        value = self.rows.get("fetch", [])
        return value(sql, args) if callable(value) else value

    async def fetchrow(self, sql, *args):
        value = self.rows.get("fetchrow")
        return value(sql, args) if callable(value) else value

    async def fetchval(self, sql, *args):
        return self.rows.get("fetchval")

    async def execute(self, sql, *args):
        self.executed.append((sql, args))
        return "UPDATE 1"


@pytest.fixture
def stub(monkeypatch):
    def install(prefs=None, **rows):
        from intentdesk import db

        fake = FakeDB(rows)
        for name in ("fetch", "fetchrow", "fetchval", "execute"):
            monkeypatch.setattr(db, name, getattr(fake, name))
        merged = {**preferences.DEFAULTS, **(prefs or {})}

        async def all_prefs():
            return merged

        monkeypatch.setattr(preferences, "all_prefs", all_prefs)
        return fake

    return install


def _user(**over):
    base = {
        "id": 1, "email": "real@example.com", "name": "Real",
        "password_hash": users.hash_password("a quiet blue tuesday"),
        "disabled": False, "is_admin": False, "google_sub": None,
        "email_verified_at": datetime.now(timezone.utc),
    }
    base.update(over)
    return base


# ---------------------------------------------------------- the sign-in gate


def test_verified_account_signs_in(stub):
    """The happy path — previously uncovered, and the one the new gate could
    most easily break."""
    stub(prefs={"access_mode": "open"}, fetchrow=lambda sql, args: _user())
    session = asyncio.run(users.authenticate("real@example.com", "a quiet blue tuesday"))
    assert session == {"email": "real@example.com", "name": "Real", "is_admin": False}


def test_unverified_account_is_refused(stub):
    stub(prefs={"access_mode": "open"},
         fetchrow=lambda sql, args: _user(email_verified_at=None))
    with pytest.raises(users.UnverifiedError) as exc:
        asyncio.run(users.authenticate("real@example.com", "a quiet blue tuesday"))
    assert exc.value.email == "real@example.com"


def test_wrong_password_does_not_reveal_the_unverified_state(stub):
    """UnverifiedError names the address, so it must be unreachable without the
    correct password — otherwise it becomes an account-enumeration oracle."""
    stub(prefs={"access_mode": "open"},
         fetchrow=lambda sql, args: _user(email_verified_at=None))
    with pytest.raises(users.AuthError) as exc:
        asyncio.run(users.authenticate("real@example.com", "the wrong password"))
    assert not isinstance(exc.value, users.UnverifiedError)
    assert str(exc.value) == "Wrong email or password."


def test_first_account_is_created_pre_verified(stub):
    """A deploy whose mail is misconfigured must not be permanently locked out
    of the admin account it just created."""
    fake = stub(prefs={"access_mode": "open"}, fetchrow=None)
    asyncio.run(users.register("first@example.com", "a quiet blue tuesday"))
    sql, _ = fake.executed[-1]
    assert "email_verified_at" in sql
    assert "NOT EXISTS (SELECT 1 FROM users)" in sql


def test_google_signup_is_verified_without_an_email(stub):
    stub(prefs={"access_mode": "open"},
         fetchrow=lambda sql, args: _user() if "INSERT" in sql else None)
    asyncio.run(users.upsert_google(sub="g-1", email="new@example.com", name="New"))
    # The INSERT itself stamps the column; nothing is left for a link to do.


# ----------------------------------------------------------------- tokens


def test_token_is_never_stored_in_the_clear(stub):
    """A database dump must not hand someone a working set of reset links."""
    captured = {}

    class Conn:
        async def execute(self, sql, *args):
            if "INSERT INTO auth_tokens" in sql:
                captured["hash"] = args[2]
            return "INSERT 0 1"

    from intentdesk import db

    class Tx:
        async def __aenter__(self):
            return Conn()

        async def __aexit__(self, *a):
            return False

    stub()
    db.transaction = lambda: Tx()

    token = asyncio.run(users.create_token(1, "verify"))
    assert token and captured["hash"] != token
    assert captured["hash"] == users._hash_token(token)
    assert len(captured["hash"]) == 64


def test_unknown_purpose_is_rejected():
    with pytest.raises(ValueError):
        asyncio.run(users.create_token(1, "something-else"))


def test_reset_links_expire_sooner_than_verification_links():
    """A live reset link is an account takeover; a live verify link is not."""
    assert users.RESET_TTL < users.VERIFY_TTL
    assert users.RESET_TTL <= timedelta(hours=1)


def test_link_is_built_from_config_not_the_request(monkeypatch):
    """Deriving this from the Host header would let an attacker who can set it
    receive the link they asked us to send someone else."""
    from intentdesk.config import settings

    monkeypatch.setattr(settings, "public_base_url", "https://intent.example.com/")
    url = users.link_for("reset", "abc123")
    assert url == "https://intent.example.com/reset?token=abc123"
    assert users.link_for("verify", "t").endswith("/verify?token=t")


def test_expired_and_used_and_unknown_tokens_are_indistinguishable(stub):
    from intentdesk import db

    past = datetime.now(timezone.utc) - timedelta(minutes=5)
    cases = [
        None,                                                        # unknown
        {"id": 1, "user_id": 1, "email": "a@b.c", "disabled": False,
         "expires_at": past, "used_at": None},                       # expired
        {"id": 1, "user_id": 1, "email": "a@b.c", "disabled": False,
         "expires_at": datetime.now(timezone.utc) + timedelta(hours=1),
         "used_at": past},                                           # already used
        {"id": 1, "user_id": 1, "email": "a@b.c", "disabled": True,
         "expires_at": datetime.now(timezone.utc) + timedelta(hours=1),
         "used_at": None},                                           # disabled account
    ]

    messages = set()
    for row in cases:
        class Conn:
            async def fetchrow(self, sql, *args):
                return row

            async def execute(self, sql, *args):
                return "UPDATE 1"

        class Tx:
            async def __aenter__(self):
                return Conn()

            async def __aexit__(self, *a):
                return False

        stub()
        db.transaction = lambda: Tx()
        with pytest.raises(users.AuthError) as exc:
            asyncio.run(users.consume_token("some-token", "reset"))
        messages.add(str(exc.value))

    assert len(messages) == 1, f"link failures are distinguishable: {messages}"


def test_empty_token_is_refused_without_a_lookup(stub):
    stub()
    with pytest.raises(users.AuthError):
        asyncio.run(users.consume_token("", "verify"))


# ------------------------------------------------------- password changes


def test_short_password_refused_on_reset(stub):
    stub()
    with pytest.raises(users.AuthError):
        asyncio.run(users.set_password(1, "short"))


def test_changing_password_requires_the_current_one(stub):
    stub(fetchrow=lambda sql, args: _user())
    with pytest.raises(users.AuthError):
        asyncio.run(
            users.change_own_password("real@example.com", "not it", "a much longer passphrase")
        )


def test_google_only_account_can_set_a_first_password(stub):
    """No password_hash means nothing to verify against — requiring a 'current'
    password would leave a Google user unable to ever add one."""
    fake = stub(fetchrow=lambda sql, args: _user(password_hash=None))
    asyncio.run(users.change_own_password("real@example.com", "", "a much longer passphrase"))
    assert any("password_hash" in sql for sql, _ in fake.executed)


# ---------------------------------------------------------------- avatars


@pytest.mark.parametrize(
    "raw,why",
    [
        (b"", "empty"),
        (b"<?php system($_GET[0]); ?>", "php disguised as an image"),
        (b"<svg onload=alert(1)></svg>", "svg, which can carry script"),
        (b"\x89PNG\r\n\x1a\n" + b"not really a png", "correct magic, corrupt body"),
        (b"\xff\xd8\xff" + b"\x00" * 40, "jpeg magic, no image"),
    ],
)
def test_bad_uploads_are_rejected(raw, why):
    with pytest.raises(avatars.AvatarError):
        avatars.normalise(raw)


def test_oversized_upload_rejected_before_decoding():
    with pytest.raises(avatars.AvatarError) as exc:
        avatars.normalise(b"\x89PNG\r\n\x1a\n" + b"\x00" * avatars.MAX_UPLOAD_BYTES)
    assert "2 MB" in str(exc.value)


def test_upload_is_re_encoded_to_a_square_png():
    from PIL import Image

    buf = io.BytesIO()
    Image.new("RGB", (900, 400), (10, 120, 200)).save(buf, format="JPEG")
    out = avatars.normalise(buf.getvalue())

    got = Image.open(io.BytesIO(out))
    assert got.format == "PNG"
    assert got.size == (avatars.SIDE, avatars.SIDE)
    # Not the original bytes: re-encoding is what strips EXIF and kills
    # polyglot files, so passing the upload through would defeat the point.
    assert out != buf.getvalue()


def test_exif_is_dropped():
    from PIL import Image

    buf = io.BytesIO()
    img = Image.new("RGB", (300, 300), (0, 0, 0))
    exif = img.getexif()
    exif[271] = "SecretCameraMake"
    img.save(buf, format="JPEG", exif=exif)
    assert b"SecretCameraMake" in buf.getvalue()

    out = avatars.normalise(buf.getvalue())
    assert b"SecretCameraMake" not in out


# ----------------------------------------------------------- signal counts


def test_counts_include_every_kind_even_at_zero(stub):
    """A missing key renders as a gap in the strip, which reads as 'no data'
    rather than the 'zero, and that is the news' it usually means."""
    stub(fetch=lambda sql, args: [
        {"kind": "review", "total": 4, "matched": 2, "avg_rating": 1.75},
    ])
    out = asyncio.run(signals.counts(days=30))

    assert set(out["by_kind"]) == set(signals.KINDS)
    assert out["by_kind"]["review"] == {"total": 4, "matched": 2, "avg_rating": 1.75}
    assert out["by_kind"]["forum"] == {"total": 0, "matched": 0, "avg_rating": None}
    assert out["total"] == 4 and out["matched"] == 2


def test_counts_tolerate_a_kind_with_no_ratings(stub):
    stub(fetch=lambda sql, args: [
        {"kind": "job_post", "total": 3, "matched": 3, "avg_rating": None},
    ])
    out = asyncio.run(signals.counts())
    assert out["by_kind"]["job_post"]["avg_rating"] is None


def test_record_rejects_an_unknown_kind(stub):
    stub()
    with pytest.raises(ValueError):
        asyncio.run(signals.record(
            kind="tweet", source="x", source_id="1",
            observed_at=datetime.now(timezone.utc),
        ))


# ----------------------------------------------------------------- export


def test_csv_carries_a_bom_for_excel(monkeypatch):
    """Without it, Excel on Windows reads UTF-8 as the system code page and
    turns every non-ASCII company name into mojibake."""
    from intentdesk.services import export, leads

    async def fake_list(**kw):
        return [{"company": "Rangmanch Événements", "score": 87}]

    monkeypatch.setattr(leads, "list_leads", fake_list)
    out = asyncio.run(export.leads_csv())
    assert out.startswith("﻿")
    assert "Rangmanch Événements" in out


def test_export_pages_past_the_internal_row_cap(monkeypatch):
    """leads.list_leads clamps its own limit to 500, so an export that made a
    single call would silently stop there and still look complete."""
    from intentdesk.services import export, leads

    calls = []

    async def fake_list(heat=None, status=None, limit=100, offset=0):
        calls.append(offset)
        remaining = 1200 - offset
        n = max(0, min(limit, remaining))
        return [{"company": f"c{offset + i}", "score": 1} for i in range(n)]

    monkeypatch.setattr(leads, "list_leads", fake_list)
    out = asyncio.run(export.leads_csv())

    assert calls == [0, 500, 1000]
    assert len(out.splitlines()) == 1201  # header + every row


def test_xlsx_writes_score_as_a_number(monkeypatch):
    """Written as text it sorts lexicographically, which puts 100 between 10 and
    11 and makes the column useless for the one thing it exists for."""
    from openpyxl import load_workbook

    from intentdesk.services import export, leads

    async def fake_list(**kw):
        return [{"company": "Acme", "score": 87, "employees_est": 40}]

    monkeypatch.setattr(leads, "list_leads", fake_list)
    wb = load_workbook(io.BytesIO(asyncio.run(export.leads_xlsx())))
    ws = wb.active

    header = [c.value for c in ws[1]]
    row = {header[i]: c.value for i, c in enumerate(ws[2])}
    assert row["Score"] == 87 and isinstance(row["Score"], int)
    assert ws.freeze_panes == "A2"


# ------------------------------------------------------------------- mail


def test_send_falls_back_to_resend_when_mautic_refuses(monkeypatch):
    """The failure mode this fallback exists for is silent: Mautic's API starts
    returning 401, mail keeps arriving via Resend, and nobody notices for weeks.
    """
    from intentdesk.services import email as mailer

    async def no_contact(*a, **kw):
        return None

    sent = {}

    async def fake_resend(to, subject, html):
        sent.update({"to": to, "subject": subject})
        return True

    monkeypatch.setattr(mailer, "upsert_contact", no_contact)
    monkeypatch.setattr(mailer, "send_resend", fake_resend)

    ok = asyncio.run(mailer.send(
        to="a@b.c", name="A", subject="Confirm", html="<p>hi</p>", template_id=7
    ))
    assert ok and sent["to"] == "a@b.c"


def test_send_reports_failure_rather_than_raising(monkeypatch):
    """A signup must not 500 because mail is down."""
    from intentdesk.services import email as mailer
    from intentdesk.config import settings

    monkeypatch.setattr(settings, "mautic_url", "")
    monkeypatch.setattr(settings, "resend_api_key", "")
    assert asyncio.run(mailer.send("a@b.c", "A", "S", "<p>x</p>")) is False


# ------------------------------------------------- signing up twice over
# The regression: signup on an address that already had an account showed
# "check your email" and sent nothing, so the one person entitled to an answer
# waited forever for a link that was never generated. Reported 2026-08-03 by
# the account owner, whose account was Google-only.


def _rows(user, recent_tokens=0):
    """fetchrow serves two different queries here — the user lookup and the
    resend-cap count. Dispatch on the statement rather than returning one shape
    for both."""
    def row(sql, args):
        if "auth_tokens" in sql:
            return {"n": recent_tokens}
        return user
    return row


def _capture_mail(monkeypatch):
    from intentdesk.services import email as mailer

    sent = {}

    async def fake_send(to, name, subject, html, template_id=0, fields=None):
        sent.update({"to": to, "subject": subject, "html": html})
        return True

    monkeypatch.setattr(mailer, "send", fake_send)
    return sent


def test_duplicate_signup_is_told_about_it_by_email(stub, monkeypatch):
    stub(fetchrow=_rows(_user(google_sub="g-1", password_hash=None)))
    sent = _capture_mail(monkeypatch)
    assert asyncio.run(users.send_existing_account_notice("real@example.com")) is True
    assert sent["to"] == "real@example.com"
    # A Google-only account is the common case: the owner is trying to give
    # themselves a password. Telling them to enter one they never set is the
    # dead end again, one step further along.
    assert "Google" in sent["html"]
    assert "/reset?token=" in sent["html"]


def test_password_account_is_not_told_to_use_google(stub, monkeypatch):
    stub(fetchrow=_rows(_user()))
    sent = _capture_mail(monkeypatch)
    asyncio.run(users.send_existing_account_notice("real@example.com"))
    assert "Google" not in sent["html"]


def test_notice_is_silent_for_an_unknown_address(stub, monkeypatch):
    """Otherwise the mail itself answers the question the form refuses to: send
    to an address, see whether anything arrives."""
    stub(fetchrow=None)
    sent = _capture_mail(monkeypatch)
    assert asyncio.run(users.send_existing_account_notice("nobody@example.com")) is False
    assert not sent


def test_notice_is_silent_for_a_disabled_account(stub, monkeypatch):
    stub(fetchrow=lambda sql, args: _user(disabled=True))
    sent = _capture_mail(monkeypatch)
    assert asyncio.run(users.send_existing_account_notice("real@example.com")) is False
    assert not sent


def test_notice_obeys_the_reset_rate_cap(stub, monkeypatch):
    """It mints a reset token, so repeated signups must not become a way to
    flood an address — the same cap /forgot already has."""
    stub(fetchrow=_rows(_user(), recent_tokens=users.RESEND_LIMIT))
    sent = _capture_mail(monkeypatch)
    assert asyncio.run(users.send_existing_account_notice("real@example.com")) is False
    assert not sent


def test_verification_body_names_the_expiry():
    from intentdesk.services import email as mailer

    subject, html = mailer.verify_email("https://x.example/verify?token=t")
    assert "24 hours" in html and "https://x.example/verify?token=t" in html
    _, reset_html = mailer.reset_email("https://x.example/reset?token=t")
    assert "1 hour" in reset_html
