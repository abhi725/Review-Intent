"""Queue export, as CSV and as a real Excel file.

CSV needs no service account, no extra OAuth scope, and Sheets imports it
directly (File → Import). The .xlsx exists because Excel is where this actually
gets opened, and a CSV opened in Excel is a worse experience than a workbook
that already has its header frozen and its columns sized.

A Google Sheets sink — writing rows into a live sheet via a service account —
stays on the roadmap. Nothing about the workflow has to wait for it.
"""

import csv
import hmac
import io
from datetime import datetime, timezone
from typing import Optional

from intentdesk.config import settings
from intentdesk.services import leads, signals


class NothingToExport(Exception):
    """No rows matched, so no file is produced.

    A header-only CSV is indistinguishable from a broken export: it downloads
    cleanly, opens cleanly, and says nothing about why it is empty. This was
    reported as "export is giving blank files" when the real state was an empty
    lead queue. Raising lets the caller say which it is.
    """

    def __init__(self, reason: str):
        self.reason = reason
        super().__init__(reason)


def _nothing_reason(status: Optional[str], heat: Optional[str]) -> str:
    filters = [f"{k}={v}" for k, v in (("status", status), ("heat", heat)) if v]
    if filters:
        return ("No leads match " + " and ".join(filters) +
                ". Clear the filters, or check the queue has been scored.")
    return ("There are no leads to export yet. The queue fills once companies "
            "have been discovered and scored — see the Sources panel.")

COLUMNS = [
    ("company", "Company"),
    ("domain", "Domain"),
    ("city", "City"),
    ("vendor", "Currently runs"),
    ("agents_est", "Agents"),
    ("score", "Score"),
    ("heat", "Heat"),
    ("status", "Status"),
    ("contact_name", "Contact"),
    ("contact_title", "Title"),
    ("contact_phone", "Phone"),
    ("contact_email", "Email"),
    ("industry", "Industry"),
    ("employees_est", "Employees"),
    ("vendor_verified", "Vendor verified"),
    ("draft_subject", "Draft subject"),
    ("draft_body", "Draft body"),
]

# Columns Excel should treat as numbers. Written as text they sort
# lexicographically, which puts 100 between 10 and 11 and makes the Score column
# useless for the one thing it exists for.
NUMERIC = {"score", "agents_est", "employees_est"}

# `leads.list_leads` clamps its own limit to 500, so a single call cannot return
# more than that no matter what is passed. Exporting therefore has to page —
# the alternative is an export that looks complete and silently is not.
PAGE = 500


async def _all_rows(status: Optional[str] = None, heat: Optional[str] = None) -> list[dict]:
    rows: list[dict] = []
    offset = 0
    while True:
        page = await leads.list_leads(heat=heat, status=status, limit=PAGE, offset=offset)
        rows.extend(page)
        if len(page) < PAGE:
            return rows
        offset += PAGE
        # A runaway guard, not a business rule. Twenty thousand leads means
        # something upstream is wrong, and a browser download is the wrong tool
        # for it either way.
        if offset >= 20_000:
            return rows


def _value(row: dict, key: str):
    v = row.get(key)
    return "" if v is None else v


async def leads_csv(status: Optional[str] = None, heat: Optional[str] = None,
                    bom: bool = True) -> str:
    """The lead queue as CSV.

    `bom` exists because the two consumers want opposite things. A browser
    download is most likely to be opened in Excel, which needs the byte-order
    mark (see below). Google Sheets' IMPORTDATA fetches the raw text and puts the
    first line straight into cells, so a BOM there becomes an invisible character
    on the front of the first column heading — a header that looks right, does not
    compare equal to "Company", and is invisible when you go looking for why.
    """
    rows = await _all_rows(status=status, heat=heat)
    if not rows:
        raise NothingToExport(_nothing_reason(status, heat))

    buf = io.StringIO()
    writer = csv.writer(buf, quoting=csv.QUOTE_MINIMAL)
    writer.writerow([label for _, label in COLUMNS])
    for row in rows:
        writer.writerow([_value(row, key) for key, _ in COLUMNS])

    # UTF-8 BOM. Excel on Windows reads a BOM-less UTF-8 CSV as the system code
    # page and turns every non-ASCII company name into mojibake. One character,
    # and it fixes the platform a download is most likely to be opened on.
    return ("﻿" if bom else "") + buf.getvalue()


async def leads_xlsx(status: Optional[str] = None, heat: Optional[str] = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    rows = await _all_rows(status=status, heat=heat)
    if not rows:
        raise NothingToExport(_nothing_reason(status, heat))

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A1A1A")
    for col, (_, label) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    for r, row in enumerate(rows, start=2):
        for c, (key, _) in enumerate(COLUMNS, start=1):
            raw = row.get(key)
            if key in NUMERIC and raw is not None:
                try:
                    raw = float(raw) if isinstance(raw, float) else int(raw)
                except (TypeError, ValueError):
                    pass
            cell = ws.cell(row=r, column=c, value="" if raw is None else raw)
            if key == "draft_body":
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Scrolling past the header on row 501 and losing track of which column is
    # which is the single most annoying thing about a long export.
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(len(rows) + 1, 1)}"

    for c, (key, label) in enumerate(COLUMNS, start=1):
        if key == "draft_body":
            width = 60
        elif key in ("draft_subject", "company", "contact_email"):
            width = 32
        else:
            longest = max((len(str(row.get(key) or "")) for row in rows), default=0)
            width = min(max(len(label) + 2, longest + 2), 28)
        ws.column_dimensions[get_column_letter(c)].width = width

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ------------------------------------------------------------- review export
#
# A second export with a different shape and a different question behind it. The
# lead export answers "who do we call next"; this one answers "what happened on
# these platforms over this period", which is the thing worth showing someone who
# is deciding whether the market is real.
#
# It reads stored rows and never collects. That is a deliberate boundary rather
# than an oversight: a date range is a cheap thing to type and the sources behind
# these rows bill per run, so wiring a fetch to a date picker would turn a typo
# into a charge.

REVIEW_COLUMNS = [
    ("period", "Period"),
    ("observed_at", "Date"),
    ("platform", "Platform"),
    ("source_site", "Source"),
    ("rating", "Rating"),
    ("category", "Complaint"),
    ("severity", "Severity"),
    ("core_complaint", "Core complaint"),
    ("switched_from", "Switched from"),
    ("switched_reason", "Switch reason"),
    ("author", "Reviewer (as published)"),
    ("country", "Country"),
    ("reviewer_name", "Reviewer (resolved)"),
    ("reviewer_title", "Reviewer title"),
    ("reviewer_company", "Reviewer company"),
    ("reviewer_confidence", "Identity confidence"),
    ("company", "Matched company"),
    ("domain", "Domain"),
    ("quote", "Title"),
    ("raw_text", "Review text"),
    ("url", "Link"),
]

REVIEW_NUMERIC = {"rating", "severity"}

SUMMARY_COLUMNS = [
    ("period", "Period"),
    ("platform", "Platform"),
    ("source_site", "Source"),
    ("reviews", "Reviews"),
    ("avg_rating", "Avg rating"),
    ("one_or_two_star", "1–2 star"),
    ("switched", "Said they switched"),
    ("matched", "Matched to a company"),
    ("top_category", "Most common complaint"),
]

SUMMARY_NUMERIC = {"reviews", "avg_rating", "one_or_two_star", "switched", "matched"}


def _review_nothing_reason(since, until, platform, source_site) -> str:
    filters = [f"{k}={v}" for k, v in (
        ("from", since.date().isoformat() if isinstance(since, datetime) else since),
        ("to", until.date().isoformat() if isinstance(until, datetime) else until),
        ("platform", platform),
        ("source", source_site),
    ) if v]
    if filters:
        return ("No stored reviews match " + " and ".join(filters) +
                ". Export reads what has already been collected — widen the range, "
                "or collect that source from the Sources panel first.")
    return ("No reviews have been collected yet. Run a source from the Sources "
            "panel; export never fetches on its own.")


def _flat(value):
    """Excel and CSV both need a scalar. Datetimes keep their type for Excel,
    which formats them as dates rather than as the string of a timestamp.

    The timezone has to go, though: openpyxl raises outright on a tz-aware
    datetime, because the .xlsx format has no way to store an offset. Converted
    to UTC first rather than simply stripped — dropping `+05:30` without shifting
    would silently move every timestamp by five and a half hours.
    """
    if value is None:
        return ""
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.astimezone(timezone.utc).replace(tzinfo=None)
    if isinstance(value, float):
        return round(value, 2)
    return value


# Minimum length for the public-export secret. It is the only thing standing
# between the lead queue and the open internet, so it is checked rather than
# trusted: a short token would be brute-forceable against a path that returns
# 404 for wrong guesses and 200 for the right one.
EXPORT_TOKEN_MIN_LEN = 24


def sheet_export_enabled() -> bool:
    """Whether the public CSV path exists at all.

    False unless a token of a sane length is configured. A public route that
    exposes leads must be switched on deliberately, and a one-character token is
    the same as no token.
    """
    token = settings.sheet_export_token or ""
    return len(token) >= EXPORT_TOKEN_MIN_LEN


def sheet_export_token_ok(candidate: str) -> bool:
    """Constant-time comparison against the configured token.

    `compare_digest` rather than `==`: the timing of a string comparison leaks how
    much of the prefix matched, which turns guessing the token from infeasible
    into a few thousand requests.
    """
    if not sheet_export_enabled():
        return False
    return hmac.compare_digest(candidate or "", settings.sheet_export_token)


# The key column for a spreadsheet sync. Google Sheets' "append or update" needs
# one column that identifies a row for its lifetime, so a re-run updates the lead
# in place instead of appending a second copy of it every time.
SHEET_KEY = "id"


async def leads_for_sheet(
    status: Optional[str] = None,
    heat: Optional[str] = None,
    limit: int = 500,
    offset: int = 0,
) -> dict:
    """Leads as flat rows for a spreadsheet sync, one page at a time.

    Paged rather than all-at-once because the caller is a scheduled HTTP client
    behind Cloudflare, which kills a request at about 100 seconds — a full export
    that times out is reported as a failed run even when the work succeeded.
    `has_more` is what lets the workflow decide to ask for the next page instead
    of guessing from the row count.

    Values are flattened to strings, numbers and booleans only. A nested object
    reaches Sheets as the text "[object Object]", which is a column of noise that
    looks like data.
    """
    rows = await leads.list_leads(heat=heat, status=status, limit=limit, offset=offset)

    out = []
    for row in rows:
        flat = {SHEET_KEY: row.get("id")}
        for key, header in COLUMNS:
            value = row.get(key)
            if isinstance(value, (dict, list)):
                # `chips` is the only one today, and a joined string is genuinely
                # more useful in a cell than a dropped column.
                value = ", ".join(str(v) for v in value) if isinstance(value, list) else ""
            flat[header] = "" if value is None else value
        created = row.get("created_at")
        flat["Created"] = created.isoformat() if hasattr(created, "isoformat") else (created or "")
        out.append(flat)

    return {
        "rows": out,
        "count": len(out),
        "offset": offset,
        "limit": limit,
        # `list_leads` clamps its own limit, so a short page is the only reliable
        # end-of-data signal.
        "has_more": len(rows) == limit,
    }


async def reviews_csv(
    since: Optional[datetime] = None,
    until: Optional[datetime] = None,
    platform: Optional[str] = None,
    source_site: Optional[str] = None,
    rating_lte: Optional[float] = None,
    group: str = "month",
) -> str:
    rows = await signals.for_export(
        since=since, until=until, platform=platform,
        source_site=source_site, rating_lte=rating_lte, group=group,
    )
    if not rows:
        raise NothingToExport(
            _review_nothing_reason(since, until, platform, source_site)
        )

    buf = io.StringIO()
    writer = csv.writer(buf, quoting=csv.QUOTE_MINIMAL)
    writer.writerow([label for _, label in REVIEW_COLUMNS])
    for row in rows:
        writer.writerow([
            # Isoformat here, not in the query: the CSV is read by tools that
            # parse ISO and by humans, and a locale-formatted date is ambiguous
            # to both.
            row["observed_at"].isoformat() if key == "observed_at" and row.get(key)
            else _flat(row.get(key))
            for key, _ in REVIEW_COLUMNS
        ])
    return "﻿" + buf.getvalue()


async def reviews_xlsx(
    since: Optional[datetime] = None,
    until: Optional[datetime] = None,
    platform: Optional[str] = None,
    source_site: Optional[str] = None,
    rating_lte: Optional[float] = None,
    group: str = "month",
) -> bytes:
    """Two sheets: Summary first, then every row.

    Summary first because that is the sheet a workbook opens on, and per-period
    counts are what someone wants before they want two thousand review bodies.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    rows = await signals.for_export(
        since=since, until=until, platform=platform,
        source_site=source_site, rating_lte=rating_lte, group=group,
    )
    if not rows:
        raise NothingToExport(
            _review_nothing_reason(since, until, platform, source_site)
        )

    summary = await signals.period_summary(
        since=since, until=until, platform=platform,
        source_site=source_site, rating_lte=rating_lte, group=group,
    )

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A1A1A")

    def write_sheet(ws, columns, data, numeric, wrap: set[str]):
        for col, (_, label) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")

        for r, row in enumerate(data, start=2):
            for c, (key, _) in enumerate(columns, start=1):
                raw = row.get(key)
                if key in numeric and raw is not None:
                    try:
                        raw = float(raw)
                    except (TypeError, ValueError):
                        pass
                cell = ws.cell(row=r, column=c, value=_flat(raw))
                if key in wrap:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                if key == "observed_at" and raw is not None:
                    cell.number_format = "yyyy-mm-dd hh:mm"

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = (
            f"A1:{get_column_letter(len(columns))}{max(len(data) + 1, 1)}"
        )
        for c, (key, label) in enumerate(columns, start=1):
            if key in wrap:
                width = 60
            else:
                longest = max((len(str(row.get(key) or "")) for row in data), default=0)
                width = min(max(len(label) + 2, longest + 2), 28)
            ws.column_dimensions[get_column_letter(c)].width = width

    ws_summary = wb.active
    ws_summary.title = "Summary"
    write_sheet(ws_summary, SUMMARY_COLUMNS, summary, SUMMARY_NUMERIC, wrap=set())

    ws_rows = wb.create_sheet("Reviews")
    write_sheet(ws_rows, REVIEW_COLUMNS, rows, REVIEW_NUMERIC,
                wrap={"raw_text", "core_complaint", "switched_reason"})

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
